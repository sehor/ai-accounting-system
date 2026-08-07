from __future__ import annotations

import json
import runpy
import sys
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "pydeps"))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
SOURCE = Path(r"C:\Users\pzr\Downloads\chkjbackup\20260806094505卡片.xls")
OUTPUT = Path(__file__).parent / "固定资产卡片-转换.xlsx"
SUMMARY = Path(__file__).parent / "fixed-asset-import-summary.json"
LEDGER_ID = "11d7a8d8-f34b-4d03-9f7f-53980b09bc88"
USER_ID = "e164807e-e122-47a0-bf4b-77980458ef25"


def clean_number(value: object, default: str = "0") -> str:
    text = str(value).strip().replace(",", "")
    return text or default


def save_workbook(frame: pd.DataFrame) -> None:
    headers = ["类别编码", "资产编码", "资产名称", "数量", "启用日期", "原值", "进项税额",
               "使用期限（月）", "残值率（%）", "期初累计折旧", "期初已折旧月数", "期初减值", "备注"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "固定资产导入"
    sheet.append(headers)
    for _, row in frame.iterrows():
        sheet.append([
            "ELECTRONIC_EQUIPMENT", str(row["编码"]).strip(), str(row["名称"]).strip(), 1,
            str(row["开始使用日期"]).strip(), Decimal(clean_number(row["资产原值"])),
            Decimal(clean_number(row["税额"])), int(Decimal(clean_number(row["预计使用期限"], "36"))),
            Decimal(clean_number(row["预计残值率"])) * 100,
            Decimal(clean_number(row["期初累计折旧"])), int(Decimal(clean_number(row["已折旧期间数"]))),
            Decimal(clean_number(row["减值准备"])), str(row["备注"]).strip(),
        ])
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    for column in sheet.columns:
        letter = column[0].column_letter
        sheet.column_dimensions[letter].width = min(36, max(14, max(len(str(cell.value or "")) for cell in column) + 2))
    sheet.freeze_panes = "A2"
    workbook.save(OUTPUT)


def main() -> None:
    frame = pd.read_excel(SOURCE, dtype=str).fillna("")
    save_workbook(frame)
    helper = runpy.run_path(str(ROOT / "test-resources" / "agent_bank_statement_e2e.py"))
    rest = helper["rest"]
    mcp = helper["McpClient"](USER_ID)

    accounts = {item["code"]: item for item in mcp.tool("list_accounts", {"ledgerId": LEDGER_ID})}
    required_codes = {"16010001001", "16010001002", "1602", "1606", "43010001008", "56020001", "53010001", "57110002"}
    missing = sorted(required_codes - accounts.keys())
    if missing:
        raise RuntimeError(f"Missing fixed-asset accounts: {missing}")

    dimension_types = mcp.tool("list_dimension_types", {"ledgerId": LEDGER_ID})
    department_type = next(item for item in dimension_types if item["code"] == "DEPARTMENT")
    department_values = mcp.tool("list_dimension_values", {
        "ledgerId": LEDGER_ID, "typeId": department_type["id"],
    })
    department = next((item for item in department_values if item["name"] == "总部"), None)
    if department is None:
        department = mcp.tool("create_dimension_value", {
            "ledgerId": LEDGER_ID,
            "typeId": department_type["id"],
            "request": {"code": "HQ", "name": "总部"},
        })

    categories = rest("GET", f"/v1/ledgers/{LEDGER_ID}/fixed-asset-categories", USER_ID)
    category = next((item for item in categories if item["code"] == "ELECTRONIC_EQUIPMENT"), None)
    if category is None:
        category = rest("POST", f"/v1/ledgers/{LEDGER_ID}/fixed-asset-categories", USER_ID, {
            "code": "ELECTRONIC_EQUIPMENT",
            "name": "电子设备",
            "usefulLifeMonths": 36,
            "residualRate": 5,
            "assetAccountId": accounts["16010001001"]["id"],
            "accumulatedDepreciationAccountId": accounts["1602"]["id"],
            "depreciationExpenseAccountId": accounts["43010001008"]["id"],
            "impairmentAccountId": accounts["1602"]["id"],
            "clearingAccountId": accounts["1606"]["id"],
            "disposalGainAccountId": accounts["53010001"]["id"],
            "disposalLossAccountId": accounts["57110002"]["id"],
        })

    page = rest("GET", f"/v1/ledgers/{LEDGER_ID}/fixed-assets?page=1&pageSize=100", USER_ID)
    existing = {item["code"]: item for item in page["data"]}
    created = 0
    for _, row in frame.iterrows():
        code = str(row["编码"]).strip()
        if code in existing:
            continue
        asset = rest("POST", f"/v1/ledgers/{LEDGER_ID}/fixed-assets", USER_ID, {
            "categoryId": category["id"],
            "code": code,
            "name": str(row["名称"]).strip(),
            "quantity": 1,
            "serviceDate": str(row["开始使用日期"]).strip(),
            "originalCost": clean_number(row["资产原值"]),
            "inputTax": clean_number(row["税额"]),
            "usefulLifeMonths": int(Decimal(clean_number(row["预计使用期限"], "36"))),
            "residualRate": str(Decimal(clean_number(row["预计残值率"])) * 100),
            "openingAccumulatedDepreciation": clean_number(row["期初累计折旧"]),
            "openingDepreciatedMonths": int(Decimal(clean_number(row["已折旧期间数"]))),
            "impairmentAmount": clean_number(row["减值准备"]),
            "departmentValueId": department["id"],
            "assetAccountId": accounts[str(row["固定资产科目"]).strip()]["id"],
            "accumulatedDepreciationAccountId": accounts[str(row["累计折旧科目"]).strip()]["id"],
            "depreciationExpenseAccountId": accounts[str(row["折旧费用分摊科目"]).strip()]["id"],
            "impairmentAccountId": accounts["1602"]["id"],
            "clearingAccountId": accounts[str(row["资产清理科目"]).strip()]["id"],
            "disposalGainAccountId": accounts["53010001"]["id"],
            "disposalLossAccountId": accounts["57110002"]["id"],
            "note": str(row["备注"]).strip() or None,
        })
        existing[code] = asset
        created += 1

    verified_page = rest("GET", f"/v1/ledgers/{LEDGER_ID}/fixed-assets?page=1&pageSize=100", USER_ID)
    verified = verified_page["data"]
    result = {
        "ledgerId": LEDGER_ID,
        "sourceAssetCount": len(frame),
        "createdAssetCount": created,
        "verifiedAssetCount": verified_page["totalItems"],
        "assetCodes": sorted(item["code"] for item in verified),
        "originalCostTotal": str(sum(Decimal(str(item["originalCost"])) for item in verified)),
        "openingAccumulatedDepreciationTotal": str(sum(Decimal(str(item["openingAccumulatedDepreciation"])) for item in verified)),
        "category": {"code": category["code"], "name": category["name"]},
        "department": {"code": department["code"], "name": department["name"]},
        "convertedWorkbook": str(OUTPUT),
    }
    SUMMARY.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
