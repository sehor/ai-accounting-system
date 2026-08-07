from __future__ import annotations

import base64
import hashlib
import json
import runpy
import sys
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "pydeps"))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
SOURCE = Path(r"C:\Users\pzr\Downloads\chkjbackup\20260806074352凭证列表_2026年第6期.xls")
OUTPUT = Path(__file__).parent / "凭证列表-2026-06-转换.xlsx"
SUMMARY = Path(__file__).parent / "voucher-import-summary.json"
LEDGER_ID = "11d7a8d8-f34b-4d03-9f7f-53980b09bc88"
USER_ID = "e164807e-e122-47a0-bf4b-77980458ef25"


def decimal(value: object) -> Decimal:
    text = str(value).strip().replace(",", "")
    return Decimal(text) if text else Decimal(0)


def save_workbook(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "凭证列表#2026年第6期"
    for row in frame.itertuples(index=False, name=None):
        sheet.append([str(value).strip() for value in row])
    for cell in sheet[3]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row != 3:
                cell.font = Font(name="Arial")
    widths = [14, 14, 34, 54, 16, 16, 14, 14]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A4"
    workbook.save(OUTPUT)


def main() -> None:
    frame = pd.read_excel(SOURCE, header=None, dtype=str).fillna("")
    save_workbook(frame)
    entries = frame.iloc[3:].copy()
    voucher_key = None
    grouped: dict[str, list[tuple[Decimal, Decimal, str]]] = {}
    for _, row in entries.iterrows():
        date = str(row.iloc[0]).strip()
        label = str(row.iloc[1]).strip()
        if date or label:
            if not date or not label:
                raise RuntimeError("Voucher date and label must both be present")
            voucher_key = f"{date}|{label}"
            grouped[voucher_key] = []
        if voucher_key is None:
            raise RuntimeError("Voucher entry without a voucher header")
        account_text = str(row.iloc[3]).strip()
        code = account_text.split(maxsplit=1)[0]
        grouped[voucher_key].append((decimal(row.iloc[4]), decimal(row.iloc[5]), code))

    unbalanced = []
    source_codes = set()
    for key, lines in grouped.items():
        debit = Decimal(0)
        credit = Decimal(0)
        for raw_debit, raw_credit, code in lines:
            source_codes.add(code)
            if raw_debit:
                (debit if raw_debit > 0 else credit)
                if raw_debit > 0:
                    debit += raw_debit
                else:
                    credit += -raw_debit
            if raw_credit:
                if raw_credit > 0:
                    credit += raw_credit
                else:
                    debit += -raw_credit
        if debit.quantize(Decimal("0.01")) != credit.quantize(Decimal("0.01")):
            unbalanced.append({"voucher": key, "debit": str(debit), "credit": str(credit)})
    if unbalanced:
        raise RuntimeError(json.dumps(unbalanced, ensure_ascii=False))

    helper = runpy.run_path(str(ROOT / "test-resources" / "agent_bank_statement_e2e.py"))
    mcp = helper["McpClient"](USER_ID)
    accounts = {item["code"] for item in mcp.tool("list_accounts", {"ledgerId": LEDGER_ID})}
    missing = sorted(source_codes - accounts)
    if missing:
        raise RuntimeError(f"Missing voucher accounts: {missing}")
    before = mcp.tool("list_vouchers", {"ledgerId": LEDGER_ID, "limit": 500, "offset": 0})
    payload = OUTPUT.read_bytes()
    key = "chkj-2026-06-" + hashlib.sha256(payload).hexdigest()[:24]
    imported = mcp.tool("import_kingdee_vouchers", {
        "ledgerId": LEDGER_ID,
        "base64Content": base64.b64encode(payload).decode(),
        "idempotencyKey": key,
    })
    after = mcp.tool("list_vouchers", {"ledgerId": LEDGER_ID, "limit": 500, "offset": 0})
    result = {
        "ledgerId": LEDGER_ID,
        "sourceVoucherCount": len(grouped),
        "sourceEntryCount": sum(len(lines) for lines in grouped.values()),
        "preImportVoucherCount": len(before),
        "mcpImportResult": imported,
        "postImportVoucherCount": len(after),
        "postedVoucherCount": sum(1 for item in after if item["status"] == "POSTED"),
        "postImportEntryCount": sum(len(item["lines"]) for item in after),
        "idempotencyKey": key,
        "convertedWorkbook": str(OUTPUT),
    }
    SUMMARY.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
