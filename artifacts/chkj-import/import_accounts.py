from __future__ import annotations

import json
import runpy
import sys
import base64
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / "pydeps"))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[2]
SOURCE = Path(r"C:\Users\pzr\Downloads\chkjbackup\20260806094517科目列表.xls")
OUTPUT = Path(__file__).parent / "科目列表-转换.xlsx"
SUMMARY = Path(__file__).parent / "account-import-summary.json"
STATE = Path(__file__).parent / "account-import-state.json"
LEDGER_ID = "11d7a8d8-f34b-4d03-9f7f-53980b09bc88"
USER_ID = "e164807e-e122-47a0-bf4b-77980458ef25"

CATEGORIES = {
    "流动资产": "ASSET",
    "非流动资产": "ASSET",
    "流动负债": "LIABILITY",
    "非流动负债": "LIABILITY",
    "所有者权益": "EQUITY",
    "成本": "COST",
    "营业收入": "REVENUE",
    "其他收益": "REVENUE",
    "营业成本及税金": "EXPENSE",
    "其他损失": "EXPENSE",
    "期间费用": "EXPENSE",
    "所得税": "EXPENSE",
    "以前年度损益调整": "EXPENSE",
}


def parent_code(code: str) -> str | None:
    if len(code) == 8:
        return code[:4]
    if len(code) in (11, 14):
        return code[:-3]
    return None


def save_source_copy(frame: pd.DataFrame) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "科目列表"
    sheet.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        sheet.append([str(value).strip() for value in row])
    for cell in sheet[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 44
    sheet.column_dimensions["C"].width = 22
    sheet.column_dimensions["D"].width = 14
    sheet.freeze_panes = "A2"
    workbook.save(OUTPUT)


def main() -> None:
    frame = pd.read_excel(SOURCE, dtype=str).fillna("")
    frame.columns = [str(value).strip() for value in frame.columns]
    for column in frame.columns:
        frame[column] = frame[column].astype(str).str.strip()
    save_source_copy(frame)

    helper = runpy.run_path(str(ROOT / "test-resources" / "agent_bank_statement_e2e.py"))
    mcp = helper["McpClient"](USER_ID)
    if STATE.exists():
        import_id = json.loads(STATE.read_text(encoding="utf-8"))["importId"]
        preview = mcp.tool("get_account_import", {"ledgerId": LEDGER_ID, "importId": import_id})
    else:
        preview = mcp.tool("preview_account_import", {
            "ledgerId": LEDGER_ID,
            "format": "KINGDEE",
            "fileName": OUTPUT.name,
            "base64Content": base64.b64encode(OUTPUT.read_bytes()).decode(),
        })
        import_id = preview["id"]
        STATE.write_text(json.dumps({"importId": import_id}, indent=2), encoding="utf-8")
    if preview["errorCount"]:
        errors = [{"rowNo": row["rowNo"], "issues": row["issues"]} for row in preview["rows"] if row["issues"]]
        raise RuntimeError(json.dumps(errors[:30], ensure_ascii=False))

    confirmed = sum(1 for row in preview["rows"] if row["confirmed"])
    for row in preview["rows"]:
        if row["confirmed"]:
            continue
        code = row["accountCode"]
        target = row["targetAccountId"]
        action = "UPDATE" if code == "5601" else "MAP" if target else "CREATE"
        mcp.tool("decide_account_import_row", {
            "ledgerId": LEDGER_ID,
            "importId": import_id,
            "rowNo": row["rowNo"],
            "decision": {
                "action": action,
                "targetAccountId": target or LEDGER_ID,
                "accountCode": code,
            },
        })
        confirmed += 1
        if confirmed % 50 == 0:
            print(f"confirmed {confirmed}/{preview['rowCount']}", flush=True)

    committed = mcp.tool("commit_account_import", {"ledgerId": LEDGER_ID, "importId": import_id})

    final_accounts = mcp.tool("list_accounts", {"ledgerId": LEDGER_ID})
    final_codes = {item["code"] for item in final_accounts}
    source_codes = set(frame["编码"])
    result = {
        "ledgerId": LEDGER_ID,
        "sourceRows": len(frame),
        "importId": import_id,
        "importStatus": committed["status"],
        "confirmedRows": sum(1 for row in committed["rows"] if row["confirmed"]),
        "finalAccountCount": len(final_accounts),
        "missingCodes": sorted(source_codes - final_codes),
        "unexpectedCodes": sorted(final_codes - source_codes),
        "convertedWorkbook": str(OUTPUT),
    }
    SUMMARY.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
